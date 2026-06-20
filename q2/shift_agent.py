"""
Shift Replacement Agent — Universitätsklinikum des Saarlandes / UKS (Homburg)

Loads real staff roster and weekly schedule from hospital_schedule_part_2.xlsx.
HR types a free-form message describing the gap; the agent:
  1. Parses role, required certifications, department, and shift type.
  2. Runs all eligibility checks (status, certs, schedule, rest, hours cap).
  3. Contacts qualified staff in priority order until someone accepts.

Usage:
  python shift_agent.py
  python shift_agent.py "Felix Haddad HOSP-1059 sick, ICU night shift 19:00-07:00 tonight, needs BLS ACLS"
"""

import os
import sys
import time
import datetime
import smtplib
import openpyxl
from email.mime.text import MIMEText
from dotenv import load_dotenv
from google import genai
from pathlib import Path
from google.genai import types

# Windows: switch console to UTF-8 so German/special characters display correctly
if sys.platform == "win32":
    os.system("chcp 65001 > nul 2>&1")
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

load_dotenv(Path(__file__).resolve().parent.parent / '.env')

# All demo notifications are forwarded to this real number/email for testing
DEMO_PHONE = os.getenv("demo_phone", "+49 15753348941")
GMAIL_ADDR = os.getenv("gmail_address", "")
GMAIL_PWD  = os.getenv("gmail_app_password", "")

API_KEY = os.getenv("gemini_api_key")
if not API_KEY:
    sys.exit("ERROR: gemini_api_key not found in .env")

client = genai.Client(api_key=API_KEY)
MODEL = "gemini-2.5-flash"

# ─── Scenario / Schedule Constants ───────────────────────────────────────────
# From the hackathon scenario sheet: decision time is Sat 20 Jun 2026 at 18:30
SCENARIO_DT   = datetime.datetime(2026, 6, 20, 18, 30)
TODAY_COL     = "Sat 06/20"
NEXT_7_DAYS   = ["Sat 06/20", "Sun 06/21", "Mon 06/22", "Tue 06/23",
                  "Wed 06/24", "Thu 06/25", "Fri 06/26"]
# Staff rested if last clock-out was before 08:30 today (not finishing today's Day shift)
RESTED_CUTOFF = datetime.datetime(2026, 6, 20, 8, 30)

XLSX = str(
    Path(__file__).resolve().parent.parent / "hackathon_problems_20260620" /
    "hackathon_problems_20260620" / "questions" / "hospital_schedule_part_2.xlsx"
)


# ─── Load Staff from Excel ────────────────────────────────────────────────────

def _simulate_response(persona: str, overtime_ok: bool, shift_pref: str) -> str:
    """Deterministic simulated response based on the persona / preferences in the roster."""
    p = (persona or "").lower()
    # Positive signals
    if any(x in p for x in ["open to last-minute", "picks up extra", "reliable"]):
        return "ACCEPT"
    # Negative signals
    if any(x in p for x in ["dislikes back-to-back", "young children",
                              "prefers predictable", "avoids overtime", "watches hours"]):
        return "DECLINE"
    # Fallback: overtime willingness + night preference
    if overtime_ok and shift_pref in ("Night", "Flexible"):
        return "ACCEPT"
    return "DECLINE"


def load_staff() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX)

    # --- Roster ---
    roster: dict[str, dict] = {}
    for row in wb["Roster"].iter_rows(min_row=2, values_only=True):
        (emp_id, fn, ln, role, dept, certs, contract,
         max_hrs, pref, ot_ok, status, persona, cin, cout, phone) = row
        if not emp_id:
            continue
        roster[emp_id] = {
            "id":             emp_id,
            "name":           f"{fn} {ln}",
            "first_name":     fn,
            "role":           role or "",
            "department":     dept or "",
            "certifications": certs or "",
            "contract":       contract or "Full-time",
            "max_hrs_week":   max_hrs or 48,
            "shift_preference": pref or "Flexible",
            "overtime_ok":    ot_ok == "Yes",
            "status":         status or "Active",
            "persona":        persona or "",
            "last_clock_out": cout,                  # datetime or str "— on shift —"
            "on_shift_now":   isinstance(cout, str), # string means still working
            "phone":          phone or "N/A",
        }

    # --- Weekly Schedule ---
    schedule: dict[str, dict] = {}
    headers = None
    for i, row in enumerate(wb["Weekly_Schedule"].iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        emp_id = row[0]
        if not emp_id:
            continue
        r = dict(zip(headers, row))
        scheduled_hrs = sum(12 for col in NEXT_7_DAYS if r.get(col) in ("D", "N"))
        schedule[emp_id] = {
            "today_slot":   r.get(TODAY_COL, "O"),
            "scheduled_hrs": scheduled_hrs,
        }

    # --- Merge and annotate ---
    staff = []
    for emp_id, r in roster.items():
        sched = schedule.get(emp_id, {"today_slot": "O", "scheduled_hrs": 0})
        r.update(sched)
        r["_resp"] = _simulate_response(r["persona"], r["overtime_ok"], r["shift_preference"])
        staff.append(r)

    return staff


print("Loading staff data from Excel...", end=" ", flush=True)
STAFF = load_staff()
print(f"done ({len(STAFF)} employees loaded)\n")

STATE = {"notified": [], "filled_by": None, "shift": {}}


# ─── Tool Implementations ─────────────────────────────────────────────────────

def find_available_staff(role: str,
                          required_certifications: str,
                          shift_type: str = "N",
                          department: str = "") -> dict:
    """
    Apply all six eligibility rules from the scenario and return a priority-sorted
    list of staff who can legally and practically cover the gap.
    """
    req_certs = {c.strip().upper() for c in required_certifications.split(",")
                 if c.strip()} if required_certifications else set()

    candidates = []
    excluded_reasons: dict[str, list[str]] = {}

    for s in STAFF:
        reasons = []

        # Rule 1: Status must be Active
        if s["status"] != "Active":
            reasons.append("on leave / inactive")

        # Rule 2: Role match
        if role.lower() not in s["role"].lower():
            reasons.append(f"wrong role ({s['role']})")

        # Rule 3: Required certifications
        if req_certs:
            staff_certs = {c.strip().upper() for c in s["certifications"].split(",") if c.strip()}
            missing = req_certs - staff_certs
            if missing:
                reasons.append(f"missing certs: {', '.join(missing)}")

        # Rule 4: Off on the target date (not already scheduled)
        if s["today_slot"] != "O":
            reasons.append(f"already scheduled ({s['today_slot']} on {TODAY_COL})")

        # Rule 5: Not currently on shift
        if s["on_shift_now"]:
            reasons.append("currently on shift")

        # Rule 6: Adequately rested
        cout = s["last_clock_out"]
        if isinstance(cout, datetime.datetime) and cout > RESTED_CUTOFF:
            reasons.append(f"not rested (clock-out {cout.strftime('%H:%M')} today)")

        # Rule 7: Won't breach weekly hours cap
        if s["scheduled_hrs"] + 12 > s["max_hrs_week"]:
            reasons.append(
                f"hours cap breached ({s['scheduled_hrs']}+12 > {s['max_hrs_week']})"
            )

        if reasons:
            excluded_reasons[s["name"]] = reasons
            continue

        # Priority scoring
        score = 0
        if s["overtime_ok"]:
            score += 10
        if shift_type == "N" and s["shift_preference"] == "Night":
            score += 8
        elif s["shift_preference"] == "Flexible":
            score += 4
        if department and department.lower() in s["department"].lower():
            score += 6  # same-dept bonus
        headroom = s["max_hrs_week"] - s["scheduled_hrs"]
        score += min(headroom // 4, 5)
        if s["contract"] in ("Per-diem", "Part-time"):
            score += 2  # cheaper/easier to call in

        candidates.append({
            "id":             s["id"],
            "name":           s["name"],
            "role":           s["role"],
            "department":     s["department"],
            "certifications": s["certifications"],
            "contract":       s["contract"],
            "overtime_ok":    s["overtime_ok"],
            "shift_preference": s["shift_preference"],
            "persona":        s["persona"],
            "scheduled_hrs":  s["scheduled_hrs"],
            "max_hrs_week":   s["max_hrs_week"],
            "hours_headroom": headroom,
            "phone":          s["phone"],
            "priority_score": score,
        })

    candidates.sort(key=lambda x: -x["priority_score"])

    # Emit structured candidate data — intercepted by the UI, hidden from terminal
    print(f"  [FOUND] {len(candidates)} eligible / {len(excluded_reasons)} excluded", flush=True)
    for i, c in enumerate(candidates[:10], 1):
        ot = "Yes" if c["overtime_ok"] else "No"
        print(
            f"  [CAND] {i}|{c['id']}|{c['name']}|{c['department']}|"
            f"{c['contract']}|{ot}|{c['priority_score']}|"
            f"{c['scheduled_hrs']}/{c['max_hrs_week']}",
            flush=True,
        )

    return {
        "found": len(candidates),
        "staff": candidates,
        "excluded_count": len(excluded_reasons),
        "eligibility_rules": [
            "1. Status = Active",
            f"2. Role contains '{role}'",
            f"3. Holds required certifications: {required_certifications or 'any'}",
            f"4. Off on {TODAY_COL} (not already scheduled)",
            "5. Not currently on shift",
            f"6. Rested — last clock-out before {RESTED_CUTOFF.strftime('%H:%M')}",
            "7. scheduled_hrs + 12 <= Max Hrs/Week",
        ],
    }


def _send_email(subject: str, body: str) -> bool:
    """Send a plain-text email via Gmail SMTP."""
    if not GMAIL_ADDR or not GMAIL_PWD:
        return False
    try:
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"]    = GMAIL_ADDR
        msg["To"]      = GMAIL_ADDR
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=10) as srv:
            srv.login(GMAIL_ADDR, GMAIL_PWD)
            srv.send_message(msg)
        return True
    except Exception as exc:
        print(f"    [email error: {exc}]")
        return False


def send_notification(emp_id: str, message: str) -> dict:
    """Send shift-cover request via SMS (demo forward) and email."""
    s = next((x for x in STAFF if x["id"] == emp_id), None)
    if not s:
        return {"success": False, "error": f"Unknown emp_id: {emp_id}"}

    print(f"  [CONTACTING] {emp_id}", flush=True)
    print(f"\n    SMS  -> {DEMO_PHONE}  (demo forward for {s['name']}, {s['id']})")
    print(f"    Mail -> {GMAIL_ADDR}")
    print(f"    Msg:   \"{message[:115]}{'...' if len(message) > 115 else ''}\"")

    sent = _send_email(
        subject=f"[UKS] Schichtvertretung — {s['name']} ({s['id']})",
        body=(
            f"SCHICHTVERTRETUNGS-ANFRAGE — UKS Homburg\n"
            f"{'='*44}\n"
            f"An:        {s['name']}  ({s['id']})\n"
            f"Rolle:     {s['role']}\n"
            f"Abteilung: {s['department']}\n"
            f"{'='*44}\n\n"
            f"{message}\n\n"
            f"{'='*44}\n"
            f"[UKS Schichtvertretungs-Agent — Demo-Modus]\n"
            f"Weitergeleitet an: {DEMO_PHONE}"
        ),
    )
    if sent:
        print(f"    [email delivered to {GMAIL_ADDR}]")

    time.sleep(0.6)

    resp = s["_resp"]
    icons  = {"ACCEPT": "[OK] ", "DECLINE": "[NO] ", "NO_RESPONSE": "[??] "}
    labels = {
        "ACCEPT":      "ACCEPTED -- will cover the shift",
        "DECLINE":     "DECLINED -- not available",
        "NO_RESPONSE": "NO RESPONSE after timeout",
    }
    print(f"    {icons.get(resp, '[??] ')} {s['name']}: {labels.get(resp, resp)}")
    print(f"  [RESPONSE] {emp_id}|{resp}", flush=True)

    STATE["notified"].append({"id": emp_id, "name": s["name"], "response": resp})
    return {
        "success":  True,
        "emp_id":   emp_id,
        "name":     s["name"],
        "response": resp,
        "label":    labels.get(resp, resp),
    }


def confirm_shift_filled(emp_id: str, department: str, shift_start: str, shift_end: str) -> dict:
    """Record the accepted assignment and generate HR confirmation."""
    s = next((x for x in STAFF if x["id"] == emp_id), None)
    if not s:
        return {"success": False, "error": "Employee not found"}

    STATE["filled_by"] = s["name"]
    STATE["shift"] = {"dept": department, "start": shift_start, "end": shift_end}
    print(f"\n    [SCHEDULE UPDATED] {s['name']} ({s['id']}) -> {department}  {shift_start}-{shift_end}")

    return {
        "success":            True,
        "assigned_employee":  s["name"],
        "employee_id":        s["id"],
        "role":               s["role"],
        "home_department":    s["department"],
        "certifications":     s["certifications"],
        "phone":              s["phone"],
        "covering_department": department,
        "shift":              f"{shift_start}-{shift_end}",
        "hr_summary": (
            f"Shift filled. {s['name']} ({s['id']}, {s['role']}) will cover "
            f"{department} from {shift_start} to {shift_end} tonight."
        ),
    }


TOOLS = {
    "find_available_staff": find_available_staff,
    "send_notification":    send_notification,
    "confirm_shift_filled": confirm_shift_filled,
}

TOOL_DECLARATIONS = [
    types.FunctionDeclaration(
        name="find_available_staff",
        description=(
            "Query the UKS staff roster and weekly schedule for employees who are eligible "
            "to cover a gap shift. Applies all eligibility rules (active status, correct role, "
            "required certifications, off today, rested, within hours cap). "
            "Returns a priority-sorted candidate list. Call this first."
        ),
        parameters={
            "type": "object",
            "properties": {
                "role": {
                    "type": "string",
                    "description": (
                        "Exact or partial role name required for the shift. "
                        "Examples: 'Registered Nurse', 'Charge Nurse', 'Physician', "
                        "'Certified Nursing Assistant', 'Nurse Practitioner'"
                    ),
                },
                "required_certifications": {
                    "type": "string",
                    "description": (
                        "Comma-separated certifications the shift requires, e.g. 'BLS, ACLS'. "
                        "Use an empty string if no specific certifications are needed."
                    ),
                },
                "shift_type": {
                    "type": "string",
                    "description": "'N' for night (19:00-07:00) or 'D' for day (07:00-19:00)",
                },
                "department": {
                    "type": "string",
                    "description": "Preferred department for priority scoring, e.g. 'ICU'. Optional.",
                },
            },
            "required": ["role", "required_certifications", "shift_type"],
        },
    ),
    types.FunctionDeclaration(
        name="send_notification",
        description=(
            "Send an automated shift-cover request to a staff member by SMS and email. "
            "Address them by first name. Write the message in German. Include the department, "
            "shift time, and contact for questions. "
            "If the response is DECLINE or NO_RESPONSE, try the next candidate on the list."
        ),
        parameters={
            "type": "object",
            "properties": {
                "emp_id": {
                    "type": "string",
                    "description": "Employee ID from find_available_staff, e.g. 'HOSP-1019'",
                },
                "message": {
                    "type": "string",
                    "description": "Shift request in German, personalized, urgent and concise",
                },
            },
            "required": ["emp_id", "message"],
        },
    ),
    types.FunctionDeclaration(
        name="confirm_shift_filled",
        description=(
            "Call once a staff member responds ACCEPT. "
            "Records the assignment in the schedule and generates the HR confirmation. "
            "This closes the case — do not contact further staff afterwards."
        ),
        parameters={
            "type": "object",
            "properties": {
                "emp_id":      {"type": "string", "description": "ID of the accepting employee"},
                "department":  {"type": "string", "description": "Department they will cover"},
                "shift_start": {"type": "string"},
                "shift_end":   {"type": "string"},
            },
            "required": ["emp_id", "department", "shift_start", "shift_end"],
        },
    ),
]

SYSTEM = """\
You are an automated shift replacement agent for Universitätsklinikum des Saarlandes (UKS Homburg).
Current date/time: Saturday 20 June 2026, 18:30.

Shift codes in the schedule: D = Day (07:00-19:00), N = Night (19:00-07:00), O = Off.

When HR reports a shift gap, follow this exact workflow:
1. Extract from the message: required role, certifications, department, and shift type (D/N).
2. Call find_available_staff to get the eligible, priority-sorted candidate list.
3. Contact the top candidate using send_notification. Write in German, addressed by first name.
4. Check the 'response' field in the tool result:
   - "ACCEPT"      -> call confirm_shift_filled, then STOP. You are done.
   - "DECLINE"     -> try the next candidate.
   - "NO_RESPONSE" -> try the next candidate.
5. If all candidates are exhausted without an ACCEPT, report this to HR with escalation options.

Rules:
- Do NOT ask follow-up questions. Infer everything from the HR message.
- Contact candidates strictly one at a time, in priority order.
- Stop the moment someone accepts and confirm_shift_filled is called.\
"""


# ─── Agent Loop ───────────────────────────────────────────────────────────────

def run_agent(hr_message: str):
    W = 68
    print(f"\n{'='*W}")
    print(f"  UKS SCHICHTVERTRETUNGS-AGENT   {SCENARIO_DT.strftime('%H:%M')}")
    print(f"  {SCENARIO_DT.strftime('%A, %d %B %Y')}  |  {MODEL}")
    print(f"{'='*W}")
    print(f"\n  HR > {hr_message}\n")
    print(f"{'-'*W}")
    print("  Agent working...\n")

    chat = client.chats.create(
        model=MODEL,
        config=types.GenerateContentConfig(
            system_instruction=SYSTEM,
            tools=[types.Tool(function_declarations=TOOL_DECLARATIONS)],
            temperature=0.1,
        ),
    )

    response = chat.send_message(hr_message)

    for _round in range(30):
        parts = response.candidates[0].content.parts
        fn_parts = [p for p in parts if getattr(p, "function_call", None)]

        if not fn_parts:
            text = "".join(p.text for p in parts if getattr(p, "text", None))
            print(f"\n{'-'*W}")
            print("  AGENT REPORT:")
            print(f"{'-'*W}")
            for line in text.strip().splitlines():
                print(f"  {line}")
            break

        tool_response_parts = []
        for p in fn_parts:
            fc = p.function_call
            args = dict(fc.args)
            arg_str = ", ".join(f"{k}={repr(v)[:42]}" for k, v in args.items())
            print(f"  [->] {fc.name}({arg_str})")

            fn = TOOLS.get(fc.name)
            result = fn(**args) if fn else {"error": f"unknown tool: {fc.name}"}

            tool_response_parts.append(
                types.Part(
                    function_response=types.FunctionResponse(name=fc.name, response=result)
                )
            )

        response = chat.send_message(tool_response_parts)

    # Footer
    print(f"\n{'='*W}")
    if STATE["filled_by"]:
        s = STATE["shift"]
        print(f"  RESULT   : SHIFT FILLED [OK]")
        print(f"  Assigned : {STATE['filled_by']}")
        print(f"  Dept     : {s.get('dept', '-')}   {s.get('start', '?')}-{s.get('end', '?')}")
        print(f"  Reached  : {len(STATE['notified'])} staff member(s)")
    else:
        print(f"  RESULT   : SHIFT NOT FILLED [FAIL]")
        print(f"  Reached  : {len(STATE['notified'])} staff member(s) -- none accepted")
    print(f"{'='*W}\n")


def main():
    if len(sys.argv) > 1:
        msg = " ".join(sys.argv[1:])
    else:
        print("UKS Shift Replacement Agent")
        print("Describe the shift gap:\n")
        msg = input("HR > ").strip()
        if not msg:
            sys.exit("No input provided.")
    run_agent(msg)


if __name__ == "__main__":
    main()
